# -*- coding: utf-8 -*-
"""
Backend para auditoría de contratos en caja fuerte - GAMAN V7.

V6 incluye:
- Peso libre, peso total y peso esperado con bolsa de seguridad.
- Bolsa de seguridad fija de 1,60 g para cálculo operativo.
- Tolerancia de novedad por peso de ±0,30 g.
- Totales por contrato y por tabla de productos.
- Archivo de salida con nombre claro y carpeta de resultados configurable.
- Inventario TOTAL incluye todos los estados presentes en el Excel.
- Lista viva de contratos pendientes por escanear/verificar.
"""

from __future__ import annotations

import csv
import json
import re
import shutil
import sys
from dataclasses import dataclass, field
from datetime import datetime, date
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


# =========================================================
# CONFIGURACIÓN GENERAL
# =========================================================
APP_VERSION = "V7"
DEVELOPER_SIGNATURE = "Desarrollado por SANTIAGO AGUDELO - GAMAN -"

if getattr(sys, "frozen", False):
    BASE_DIR = Path(sys.executable).resolve().parent
else:
    BASE_DIR = Path(__file__).resolve().parent

APP_DIR = BASE_DIR / "Resultados"
CONFIG_FILE = APP_DIR / "config_caja_fuerte.json"

MODO_TOTAL = "TOTAL"
MODO_PARCIAL = "PARCIAL"

# La bolsa no pertenece al contrato; se usa solo para auditoría física.
PESO_BOLSA_SEGURIDAD_GR = 1.60

# Si el auditor digita un peso real con bolsa y la diferencia neta supera esto, queda como novedad.
PESO_TOLERANCIA_GR = 0.30

# El estado del contrato se conserva para auditoría, pero V6 ya no excluye contratos por estado.
ESTADO_ACTIVO = "ACTIVO"

# Encabezados de auditoría que se agregan al Excel de salida.
AUDIT_COLUMNS = [
    "AUDITORIA_ESTADO",
    "AUDITORIA_FECHA",
    "AUDITORIA_MODO",
    "AUDITORIA_USUARIO",
    "AUDITORIA_RESULTADO",
    "AUDITORIA_NOVEDAD_TIPO",
    "AUDITORIA_CANTIDAD_ESPERADA",
    "AUDITORIA_CANTIDAD_EVIDENCIADA",
    "AUDITORIA_PESO_LIBRE_ESPERADO",
    "AUDITORIA_PESO_TOTAL_ESPERADO",
    "AUDITORIA_PESO_BOLSA_SEGURIDAD",
    "AUDITORIA_PESO_CON_BOLSA_ESPERADO",
    "AUDITORIA_PESO_BRUTO_BOLSA_EVIDENCIADO",
    "AUDITORIA_PESO_NETO_CALCULADO",
    "AUDITORIA_DIFERENCIA_PESO_TOTAL",
    "AUDITORIA_BOLSA_VALIDACION_VISUAL",
    "AUDITORIA_OBSERVACION",
    "AUDITORIA_RAW_ESCANEO",
]

# Posibles nombres de columnas en diferentes reportes.
COLUMN_CANDIDATES = {
    "contrato": ["Contrato Secuencia", "Contrato", "Secuencia Contrato", "Numero Contrato", "Número Contrato"],
    "estado": ["Contrato Estado", "Estado", "Estado Contrato"],
    "joyeria": ["Joyeria", "Joyería"],
    "ciudad": ["Ciudad"],
    "cliente": ["Cliente", "Cliente Nombre", "Nombre Cliente", "Nombre"],
    "documento": ["Documento Cliente", "Documento", "C.C.", "Cedula", "Cédula"],
    "valor": ["Contrato Valor", "Valor", "Valor Contrato", "Valor Total"],
    "fecha_inicio": ["Fecha Inicio", "Fecha Contrato", "Fecha"],
    "fecha_operacion": ["Fecha Operacion", "Fecha Operación", "Fecha Creacion", "Fecha Creación"],
    "gramos_contrato_estimados": ["Contrato Gramos Estimados", "Gramos Estimados", "Peso Libre", "Gramos Libres"],
    "gramos_contrato_totales": ["Contrato Gramos Totales", "Gramos Totales", "Peso Total"],
    "numero_producto": ["Numero Producto", "Número Producto", "Cant. Articulos", "Cant. Artículos"],
    "familia": ["Familia", "Producto", "Tipo Producto"],
    "calidad": ["Producto Calidad", "Calidad"],
    "descripcion": ["Producto Descripcion", "Producto Descripción", "Descripcion", "Descripción", "Detalle"],
    "producto_peso_estimado": ["Producto Peso Estimado", "Peso Estimado Producto"],
    "producto_peso_total": ["Producto Peso Total", "Peso Total Producto"],
    "producto_valor_compra": ["Producto Valor Compra", "Valor Compra Producto"],
    "tipo_articulo": ["Tipo Articulo", "Tipo Artículo"],
}


# =========================================================
# FUNCIONES UTILITARIAS
# =========================================================
def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_key(value: Any) -> str:
    s = normalize_text(value).lower()
    replacements = {
        "á": "a", "é": "e", "í": "i", "ó": "o", "ú": "u", "ñ": "n",
    }
    for a, b in replacements.items():
        s = s.replace(a, b)
    s = re.sub(r"\s+", " ", s)
    return s


def sanitize_filename(value: str) -> str:
    value = normalize_text(value) or "SIN_JOYERIA"
    value = re.sub(r"[^A-Za-z0-9_-]+", "_", value)
    value = re.sub(r"_+", "_", value).strip("_")
    return value or "SIN_JOYERIA"


def normalize_contract(raw: Any) -> Optional[str]:
    if raw is None:
        return None
    s = str(raw).strip()
    if not s:
        return None
    
    # Extract all digit sequences from the input
    parts = re.findall(r"\d+", s)
    
    # REGLA PARA CONTRATOS VIEJOS (5 dígitos exactos)
    # Si el valor escaneado contiene exactamente 5 dígitos, interpretarlo como contrato viejo
    # Formato SINNUT viejo: 434-99- + número a 6 dígitos
    # Ejemplo: 35265 -> 434-99-035265
    if len(parts) == 1 and len(parts[0]) == 5:
        digits = parts[0]
        return f"434-99-{digits.zfill(6)}"
    
    # Lógica actual para contratos nuevos
    # Formatos soportados:
    # - 434-01-003162
    # - 43401003162
    # - 434 01 003162
    # - 434'01'003162
    if len(parts) >= 3:
        a, b, c = parts[0], parts[1], parts[2]
        return f"{a.zfill(3)}-{b.zfill(2)}-{c.zfill(6)}"
    if len(parts) == 1:
        digits = parts[0]
        if len(digits) >= 5:
            a = digits[:3].zfill(3)
            b = digits[3:5].zfill(2)
            c = digits[5:].zfill(6)
            return f"{a}-{b}-{c}"
    return None


def parse_float(value: Any) -> Optional[float]:
    if value is None or value == "":
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    s = re.sub(r"[^\d,.-]", "", s)
    if not s:
        return None
    if "," in s and "." in s:
        # Formato colombiano: 1.234,56
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def parse_int(value: Any) -> Optional[int]:
    if value is None or value == "":
        return None
    if isinstance(value, int):
        return value
    if isinstance(value, float):
        return int(round(value))
    s = re.sub(r"[^\d-]", "", str(value))
    if not s:
        return None
    try:
        return int(s)
    except ValueError:
        return None


def parse_date(value: Any) -> Optional[date]:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    formats = [
        "%d/%m/%Y",
        "%d-%m-%Y",
        "%Y-%m-%d",
        "%Y-%m-%d %H:%M:%S",
        "%d/%m/%Y %H:%M:%S",
        "%m/%d/%Y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    match = re.search(r"(\d{1,2})[/-](\d{1,2})[/-](\d{4})", s)
    if match:
        d, m, y = match.groups()
        try:
            return date(int(y), int(m), int(d))
        except ValueError:
            return None
    match = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if match:
        y, m, d = match.groups()
        try:
            return date(int(y), int(m), int(d))
        except ValueError:
            return None
    return None


def format_money(value: Any) -> str:
    n = parse_float(value)
    if n is None:
        return normalize_text(value)
    return "$ " + f"{int(round(n)):,}".replace(",", ".")


def format_document(value: Any) -> str:
    s = normalize_text(value)
    digits = re.sub(r"\D", "", s)
    if not digits:
        return s
    return f"{int(digits):,}".replace(",", ".")


def format_decimal(value: Any, digits: int = 2) -> str:
    n = parse_float(value)
    if n is None:
        return ""
    return f"{n:.{digits}f}".replace(".", ",")


def format_decimal_dot(value: Any, digits: int = 2) -> str:
    n = parse_float(value)
    if n is None:
        return ""
    return f"{n:.{digits}f}"


def parse_user_date(value: str) -> Optional[date]:
    value = normalize_text(value)
    if not value:
        return None
    return parse_date(value)


def now_stamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def build_out_dir(run_stamp: str, output_base: str | Path | None = None) -> Path:
    base = Path(str(output_base).strip().strip('"')) if output_base else APP_DIR
    base.mkdir(parents=True, exist_ok=True)
    out_dir = base / f"auditoria_caja_fuerte_{run_stamp}"
    out_dir.mkdir(parents=True, exist_ok=True)
    return out_dir


def load_config() -> dict[str, Any]:
    try:
        if CONFIG_FILE.exists():
            return json.loads(CONFIG_FILE.read_text(encoding="utf-8"))
    except Exception:
        pass
    return {}


def save_config(data: dict[str, Any]) -> None:
    APP_DIR.mkdir(parents=True, exist_ok=True)
    CONFIG_FILE.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")


def get_last_run_text() -> str:
    cfg = load_config()
    return normalize_text(cfg.get("last_run", ""))


def get_last_run_date() -> Optional[date]:
    cfg = load_config()
    return parse_date(cfg.get("last_run", ""))


def get_default_output_dir() -> str:
    cfg = load_config()
    return normalize_text(cfg.get("output_dir", str(APP_DIR))) or str(APP_DIR)


# =========================================================
# ESTRUCTURAS DE DATOS
# =========================================================
@dataclass
class ContractRecord:
    contrato: str
    rows: list[int]
    activo: bool
    en_alcance: bool
    orden: int = 0
    siguiente_contrato: str = ""
    motivo_fuera_alcance: str = ""
    estado: str = ""
    joyeria: str = ""
    ciudad: str = ""
    cliente: str = ""
    documento: str = ""
    valor: Any = None
    fecha_inicio: str = ""
    fecha_operacion: str = ""
    fecha_base: Optional[date] = None
    cantidad_esperada: int = 0
    peso_libre_esperado: Optional[float] = None
    peso_total_esperado: Optional[float] = None
    peso_bolsa_seguridad: float = PESO_BOLSA_SEGURIDAD_GR
    peso_con_bolsa_esperado: Optional[float] = None
    peso_productos_libre: Optional[float] = None
    peso_productos_total: Optional[float] = None
    valor_productos_total: Optional[float] = None
    productos: list[dict[str, Any]] = field(default_factory=list)
    audit_estado: str = ""
    audit_fecha: str = ""
    audit_novedad_tipo: str = ""
    audit_observacion: str = ""

    def to_gui_dict(self) -> dict[str, Any]:
        return {
            "contrato": self.contrato,
            "estado": self.estado,
            "joyeria": self.joyeria,
            "ciudad": self.ciudad,
            "cliente": self.cliente,
            "documento": format_document(self.documento),
            "documento_raw": self.documento,
            "valor": format_money(self.valor),
            "fecha_inicio": self.fecha_inicio,
            "fecha_operacion": self.fecha_operacion,
            "cantidad_esperada": self.cantidad_esperada,
            "peso_libre_esperado": format_decimal(self.peso_libre_esperado),
            "peso_total_esperado": format_decimal(self.peso_total_esperado),
            "peso_bolsa_seguridad": format_decimal(self.peso_bolsa_seguridad),
            "peso_con_bolsa_esperado": format_decimal(self.peso_con_bolsa_esperado),
            "peso_productos_libre": format_decimal(self.peso_productos_libre),
            "peso_productos_total": format_decimal(self.peso_productos_total),
            "valor_productos_total": format_money(self.valor_productos_total),
            "productos": self.productos,
            "audit_estado": self.audit_estado,
            "audit_fecha": self.audit_fecha,
            "audit_novedad_tipo": self.audit_novedad_tipo,
            "audit_observacion": self.audit_observacion,
            "activo": self.activo,
            "en_alcance": self.en_alcance,
            "motivo_fuera_alcance": self.motivo_fuera_alcance,
            "siguiente_contrato": self.siguiente_contrato,
        }


@dataclass
class ScanResult:
    status: str
    message: str
    contrato: Optional[str] = None
    record: Optional[dict[str, Any]] = None


# =========================================================
# BACKEND PRINCIPAL
# =========================================================
class CajaFuerteBackend:
    def __init__(self) -> None:
        self.excel_path: Optional[Path] = None
        self.out_dir: Optional[Path] = None
        self.out_xlsx: Optional[Path] = None
        self.log_csv: Optional[Path] = None
        self.run_stamp: str = ""
        self.modo: str = MODO_TOTAL
        self.fecha_desde: Optional[date] = None
        self.usuario: str = ""
        self.output_base: Optional[Path] = None

        self.wb = None
        self.ws = None
        self.header_row: int = 1
        self.cols: dict[str, Optional[int]] = {}
        self.audit_cols: dict[str, int] = {}

        self.all_contracts: dict[str, ContractRecord] = {}
        self.target_contracts: dict[str, ContractRecord] = {}
        self.no_encontrados: dict[str, dict[str, str]] = {}
        self.audit_log: list[list[Any]] = []

    # ---------------------------
    # Carga e indexación
    # ---------------------------
    def start(
        self,
        excel_path: str | Path,
        modo: str,
        fecha_desde: Optional[date] = None,
        usuario: str = "",
        output_dir: str | Path | None = None,
    ) -> dict[str, Any]:
        self.excel_path = Path(str(excel_path).strip().strip('"'))
        if not self.excel_path.exists():
            raise FileNotFoundError(f"No existe el archivo: {self.excel_path}")

        self.modo = modo.upper().strip()
        if self.modo not in (MODO_TOTAL, MODO_PARCIAL):
            raise ValueError("El modo debe ser TOTAL o PARCIAL.")

        self.fecha_desde = fecha_desde
        if self.modo == MODO_PARCIAL and not self.fecha_desde:
            raise ValueError("Para auditoría PARCIAL debes indicar una fecha desde.")

        self.usuario = normalize_text(usuario) or "AUDITOR"
        self.run_stamp = now_stamp()
        self.out_dir = build_out_dir(self.run_stamp, output_dir)
        self.output_base = self.out_dir.parent
        self.log_csv = self.out_dir / "log_escaneos_caja_fuerte.csv"
        self.audit_log.clear()
        self.no_encontrados.clear()

        temp_xlsx = self.out_dir / f"_trabajo_auditoria_{self.run_stamp}.xlsx"
        shutil.copy2(self.excel_path, temp_xlsx)

        self.wb = load_workbook(temp_xlsx)
        self.ws = self.wb.active

        self.header_row = self._detect_header_row()
        self.cols = self._detect_columns()
        if not self.cols.get("contrato"):
            raise ValueError("No encontré la columna de contrato. Debe existir 'Contrato Secuencia' o equivalente.")

        self.audit_cols = self._ensure_audit_columns()
        self._build_contract_indexes()
        self._assign_next_contracts()

        joyeria = self._detect_output_joyeria()
        stamp_human = datetime.now().strftime("%Y-%m-%d_%H-%M")
        final_name = f"AUDITORIA_CAJA_FUERTE_{self.modo}_{sanitize_filename(joyeria)}_{stamp_human}.xlsx"
        self.out_xlsx = self.out_dir / final_name
        self._open_log_csv()
        self.save()
        try:
            if temp_xlsx.exists() and temp_xlsx != self.out_xlsx:
                temp_xlsx.unlink()
        except OSError:
            pass

        return self.get_progress()

    def _detect_output_joyeria(self) -> str:
        for record in self.target_contracts.values():
            if record.joyeria:
                return record.joyeria
        for record in self.all_contracts.values():
            if record.joyeria:
                return record.joyeria
        return "SIN_JOYERIA"

    def _detect_header_row(self) -> int:
        targets = {normalize_key(x) for x in COLUMN_CANDIDATES["contrato"]}
        max_rows = min(20, self.ws.max_row)
        for row in range(1, max_rows + 1):
            for col in range(1, self.ws.max_column + 1):
                if normalize_key(self.ws.cell(row=row, column=col).value) in targets:
                    return row
        return 1

    def _detect_columns(self) -> dict[str, Optional[int]]:
        header_map: dict[str, int] = {}
        for col in range(1, self.ws.max_column + 1):
            value = self.ws.cell(row=self.header_row, column=col).value
            key = normalize_key(value)
            if key:
                header_map[key] = col

        cols: dict[str, Optional[int]] = {}
        for logical_name, candidates in COLUMN_CANDIDATES.items():
            found = None
            for candidate in candidates:
                key = normalize_key(candidate)
                if key in header_map:
                    found = header_map[key]
                    break
            cols[logical_name] = found
        return cols

    def _ensure_audit_columns(self) -> dict[str, int]:
        existing: dict[str, int] = {}
        for col in range(1, self.ws.max_column + 1):
            key = normalize_key(self.ws.cell(row=self.header_row, column=col).value)
            if key:
                existing[key] = col

        audit_cols: dict[str, int] = {}
        for header in AUDIT_COLUMNS:
            key = normalize_key(header)
            if key in existing:
                audit_cols[header] = existing[key]
            else:
                new_col = self.ws.max_column + 1
                self.ws.cell(row=self.header_row, column=new_col).value = header
                self.ws.cell(row=self.header_row, column=new_col).font = Font(bold=True, color="FFFFFF")
                self.ws.cell(row=self.header_row, column=new_col).fill = PatternFill("solid", fgColor="1F4E78")
                self.ws.cell(row=self.header_row, column=new_col).alignment = Alignment(horizontal="center")
                audit_cols[header] = new_col
        return audit_cols

    def _cell(self, row: int, logical_col: str) -> Any:
        col = self.cols.get(logical_col)
        if not col:
            return None
        return self.ws.cell(row=row, column=col).value

    def _audit_cell(self, row: int, header: str) -> Any:
        col = self.audit_cols.get(header)
        if not col:
            return None
        return self.ws.cell(row=row, column=col).value

    def _first_non_empty(self, rows: list[int], logical_col: str) -> Any:
        for row in rows:
            value = self._cell(row, logical_col)
            if value not in (None, ""):
                return value
        return None

    def _sum_float(self, rows: list[int], logical_col: str) -> Optional[float]:
        total = 0.0
        found = False
        for row in rows:
            value = parse_float(self._cell(row, logical_col))
            if value is not None:
                total += value
                found = True
        return total if found else None

    def _count_expected_items(self, rows: list[int]) -> int:
        col = self.cols.get("numero_producto")
        if col:
            values = set()
            for row in rows:
                value = self.ws.cell(row=row, column=col).value
                if value not in (None, ""):
                    values.add(str(value).strip())
            if values:
                return len(values)
        return len(rows)

    def _build_contract_indexes(self) -> None:
        grouped: dict[str, list[int]] = {}
        contrato_col = self.cols["contrato"]

        for row in range(self.header_row + 1, self.ws.max_row + 1):
            norm = normalize_contract(self.ws.cell(row=row, column=contrato_col).value)
            if not norm:
                continue
            grouped.setdefault(norm, []).append(row)

        today = date.today()
        self.all_contracts.clear()
        self.target_contracts.clear()

        for orden, (contrato, rows) in enumerate(grouped.items(), start=1):
            estado = normalize_text(self._first_non_empty(rows, "estado"))
            activo = normalize_key(estado).upper() == ESTADO_ACTIVO

            fecha_inicio_raw = self._first_non_empty(rows, "fecha_inicio")
            fecha_operacion_raw = self._first_non_empty(rows, "fecha_operacion")
            fecha_base = parse_date(fecha_operacion_raw) or parse_date(fecha_inicio_raw)

            peso_contrato_libre = parse_float(self._first_non_empty(rows, "gramos_contrato_estimados"))
            peso_contrato_total = parse_float(self._first_non_empty(rows, "gramos_contrato_totales"))
            peso_productos_libre = self._sum_float(rows, "producto_peso_estimado")
            peso_productos_total = self._sum_float(rows, "producto_peso_total")
            valor_productos_total = self._sum_float(rows, "producto_valor_compra")

            peso_libre_esperado = peso_contrato_libre if peso_contrato_libre is not None else peso_productos_libre
            peso_total_esperado = peso_contrato_total if peso_contrato_total is not None else peso_productos_total
            peso_con_bolsa_esperado = None
            if peso_total_esperado is not None:
                peso_con_bolsa_esperado = round(peso_total_esperado + PESO_BOLSA_SEGURIDAD_GR, 2)

            productos: list[dict[str, Any]] = []
            for row in rows:
                productos.append({
                    "numero": normalize_text(self._cell(row, "numero_producto")),
                    "familia": normalize_text(self._cell(row, "familia")),
                    "calidad": normalize_text(self._cell(row, "calidad")),
                    "valor_compra": format_money(self._cell(row, "producto_valor_compra")),
                    "descripcion": normalize_text(self._cell(row, "descripcion")),
                    "peso_estimado": format_decimal(self._cell(row, "producto_peso_estimado")),
                    "peso_total": format_decimal(self._cell(row, "producto_peso_total")),
                    "tipo_articulo": normalize_text(self._cell(row, "tipo_articulo")),
                })

            audit_estado = normalize_text(self._audit_cell(rows[0], "AUDITORIA_ESTADO"))
            audit_fecha = normalize_text(self._audit_cell(rows[0], "AUDITORIA_FECHA"))
            audit_novedad_tipo = normalize_text(self._audit_cell(rows[0], "AUDITORIA_NOVEDAD_TIPO"))
            audit_observacion = normalize_text(self._audit_cell(rows[0], "AUDITORIA_OBSERVACION"))

            # V6: el estado NO excluye contratos del alcance.
            # El auditor necesita ver también resolucionados, vencidos, retirados, pagados, etc.
            # TOTAL = todos los contratos presentes en el Excel.
            # PARCIAL = contratos dentro del rango de fecha seleccionado, conservando cualquier estado.
            en_alcance = False
            motivo = ""
            if self.modo == MODO_TOTAL:
                en_alcance = True
            else:
                if fecha_base is None:
                    en_alcance = True
                    motivo = "Incluido por seguridad: contrato sin fecha base."
                elif self.fecha_desde <= fecha_base <= today:
                    en_alcance = True
                else:
                    motivo = f"Fuera del rango parcial: fecha {fecha_base.strftime('%d/%m/%Y')}"

            record = ContractRecord(
                contrato=contrato,
                rows=rows,
                activo=activo,
                en_alcance=en_alcance,
                orden=orden,
                motivo_fuera_alcance=motivo,
                estado=estado,
                joyeria=normalize_text(self._first_non_empty(rows, "joyeria")),
                ciudad=normalize_text(self._first_non_empty(rows, "ciudad")),
                cliente=normalize_text(self._first_non_empty(rows, "cliente")),
                documento=normalize_text(self._first_non_empty(rows, "documento")),
                valor=self._first_non_empty(rows, "valor"),
                fecha_inicio=normalize_text(fecha_inicio_raw),
                fecha_operacion=normalize_text(fecha_operacion_raw),
                fecha_base=fecha_base,
                cantidad_esperada=self._count_expected_items(rows),
                peso_libre_esperado=peso_libre_esperado,
                peso_total_esperado=peso_total_esperado,
                peso_bolsa_seguridad=PESO_BOLSA_SEGURIDAD_GR,
                peso_con_bolsa_esperado=peso_con_bolsa_esperado,
                peso_productos_libre=peso_productos_libre,
                peso_productos_total=peso_productos_total,
                valor_productos_total=valor_productos_total,
                productos=productos,
                audit_estado=audit_estado,
                audit_fecha=audit_fecha,
                audit_novedad_tipo=audit_novedad_tipo,
                audit_observacion=audit_observacion,
            )

            self.all_contracts[contrato] = record
            if record.en_alcance:
                self.target_contracts[contrato] = record

    def _assign_next_contracts(self) -> None:
        ordered = sorted(self.target_contracts.values(), key=lambda r: min(r.rows) if r.rows else r.orden)
        for idx, record in enumerate(ordered):
            record.siguiente_contrato = ordered[idx + 1].contrato if idx + 1 < len(ordered) else "FIN"

    # ---------------------------
    # Escaneo y registro
    # ---------------------------
    def scan_contract(self, raw_scan: str) -> ScanResult:
        raw_scan = normalize_text(raw_scan)
        norm = normalize_contract(raw_scan)
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not norm:
            self._append_log([ts, raw_scan, "", "INVALIDO", "No se pudo normalizar la lectura"])
            return ScanResult("INVALIDO", f"Lectura inválida: {raw_scan}")

        record = self.target_contracts.get(norm)
        if record:
            if record.audit_estado:
                msg = f"Contrato ya auditado como {record.audit_estado}. Puedes revisar o sobrescribir si fue un error."
                status = "YA_AUDITADO"
            else:
                msg = "Contrato encontrado en el alcance de auditoría."
                status = "ENCONTRADO"
            self._append_log([ts, raw_scan, norm, status, msg])
            return ScanResult(status, msg, norm, record.to_gui_dict())

        any_record = self.all_contracts.get(norm)
        if any_record:
            msg = any_record.motivo_fuera_alcance or "El contrato existe, pero está fuera del alcance seleccionado."
            status = "FUERA_ALCANCE"
            self._append_log([ts, raw_scan, norm, status, msg])
            return ScanResult(status, msg, norm, any_record.to_gui_dict())

        msg = "El contrato no existe en el Excel cargado."
        self.no_encontrados[norm] = {"raw": raw_scan, "timestamp": ts, "detalle": msg}
        self._append_log([ts, raw_scan, norm, "NO_EXISTE", msg])
        return ScanResult("NO_EXISTE", msg, norm, None)

    def register_verification(
        self,
        contrato: str,
        raw_scan: str = "",
        cantidad_evidenciada: Any = None,
        peso_bruto_bolsa_evidenciado: Any = None,
        novedad_tipo: str = "OK",
        observacion: str = "",
    ) -> dict[str, Any]:
        norm = normalize_contract(contrato)
        if not norm or norm not in self.target_contracts:
            raise ValueError("No hay contrato válido para registrar en el alcance actual.")

        record = self.target_contracts[norm]
        ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cantidad = parse_int(cantidad_evidenciada)
        peso_bruto = parse_float(peso_bruto_bolsa_evidenciado)
        novedad_tipo = normalize_text(novedad_tipo).upper() or "OK"
        observacion = normalize_text(observacion)

        novedades: list[str] = []
        if novedad_tipo != "OK":
            novedades.append(novedad_tipo)

        if cantidad is not None and cantidad != record.cantidad_esperada:
            novedades.append("DIFERENCIA_CANTIDAD")

        peso_neto_calculado = None
        diff_peso_total = None
        if peso_bruto is not None:
            peso_neto_calculado = round(peso_bruto - PESO_BOLSA_SEGURIDAD_GR, 3)
            if record.peso_total_esperado is not None:
                diff_peso_total = round(peso_neto_calculado - record.peso_total_esperado, 3)
                if abs(diff_peso_total) > PESO_TOLERANCIA_GR:
                    novedades.append("DIFERENCIA_PESO_TOTAL")

        # La bolsa se valida visualmente; al marcar OK se toma como correcta.
        bolsa_validacion_visual = "OK_VISUAL" if novedad_tipo == "OK" else "NOVEDAD_VISUAL"

        clean_novedades: list[str] = []
        for item in novedades:
            if item and item not in clean_novedades:
                clean_novedades.append(item)

        resultado = "OK" if not clean_novedades else "NOVEDAD"
        novedad_final = "OK" if resultado == "OK" else " | ".join(clean_novedades)

        if resultado == "NOVEDAD" and not observacion:
            observacion = "Novedad registrada por diferencia detectada durante la auditoría."

        fill_ok = PatternFill("solid", fgColor="C6EFCE")
        fill_nov = PatternFill("solid", fgColor="FFF2CC")
        fill = fill_ok if resultado == "OK" else fill_nov

        for row in record.rows:
            values = {
                "AUDITORIA_ESTADO": resultado,
                "AUDITORIA_FECHA": ts,
                "AUDITORIA_MODO": self.modo,
                "AUDITORIA_USUARIO": self.usuario,
                "AUDITORIA_RESULTADO": resultado,
                "AUDITORIA_NOVEDAD_TIPO": novedad_final,
                "AUDITORIA_CANTIDAD_ESPERADA": record.cantidad_esperada,
                "AUDITORIA_CANTIDAD_EVIDENCIADA": cantidad,
                "AUDITORIA_PESO_LIBRE_ESPERADO": record.peso_libre_esperado,
                "AUDITORIA_PESO_TOTAL_ESPERADO": record.peso_total_esperado,
                "AUDITORIA_PESO_BOLSA_SEGURIDAD": PESO_BOLSA_SEGURIDAD_GR,
                "AUDITORIA_PESO_CON_BOLSA_ESPERADO": record.peso_con_bolsa_esperado,
                "AUDITORIA_PESO_BRUTO_BOLSA_EVIDENCIADO": peso_bruto,
                "AUDITORIA_PESO_NETO_CALCULADO": peso_neto_calculado,
                "AUDITORIA_DIFERENCIA_PESO_TOTAL": diff_peso_total,
                "AUDITORIA_BOLSA_VALIDACION_VISUAL": bolsa_validacion_visual,
                "AUDITORIA_OBSERVACION": observacion,
                "AUDITORIA_RAW_ESCANEO": raw_scan,
            }
            for header, value in values.items():
                cell = self.ws.cell(row=row, column=self.audit_cols[header])
                cell.value = value
                cell.fill = fill

        record.audit_estado = resultado
        record.audit_fecha = ts
        record.audit_novedad_tipo = novedad_final
        record.audit_observacion = observacion

        detalle = (
            f"{resultado} | {novedad_final} | Cantidad {cantidad}/{record.cantidad_esperada} | "
            f"Peso total esperado {record.peso_total_esperado} | Peso con bolsa esperado {record.peso_con_bolsa_esperado} | "
            f"Peso bruto evidenciado {peso_bruto} | Neto calculado {peso_neto_calculado} | Diff {diff_peso_total}"
        )
        self._append_log([ts, raw_scan, norm, resultado, detalle])
        self.save()

        return {
            "resultado": resultado,
            "novedad_tipo": novedad_final,
            "observacion": observacion,
            "diff_peso_total": diff_peso_total,
            "peso_neto_calculado": peso_neto_calculado,
            "progress": self.get_progress(),
        }

    # ---------------------------
    # Progreso y finalización
    # ---------------------------
    def is_verified(self, record: ContractRecord) -> bool:
        return normalize_key(record.audit_estado) in {"ok", "novedad"}

    def get_progress(self) -> dict[str, Any]:
        total = len(self.target_contracts)
        verificados = sum(1 for r in self.target_contracts.values() if self.is_verified(r))
        novedades = sum(1 for r in self.target_contracts.values() if normalize_key(r.audit_estado) == "novedad")
        faltantes = max(total - verificados, 0)
        no_encontrados = len(self.no_encontrados)
        return {
            "modo": self.modo,
            "fecha_desde": self.fecha_desde.strftime("%d/%m/%Y") if self.fecha_desde else "",
            "total_objetivo": total,
            "verificados": verificados,
            "novedades": novedades,
            "faltantes": faltantes,
            "no_encontrados": no_encontrados,
            "archivo_salida": str(self.out_xlsx) if self.out_xlsx else "",
            "carpeta_salida": str(self.out_dir) if self.out_dir else "",
        }

    def get_pending_contracts(self, limit: int = 300) -> list[dict[str, Any]]:
        """Devuelve contratos del alcance que aún no tienen OK/NOVEDAD.

        Se usa en la GUI para que el auditor vea, antes de finalizar,
        cuáles contratos siguen pendientes de encontrar o escanear.
        """
        pending: list[ContractRecord] = [
            record for record in self.target_contracts.values()
            if not self.is_verified(record)
        ]
        pending.sort(key=lambda r: min(r.rows) if r.rows else r.orden)

        rows: list[dict[str, Any]] = []
        for record in pending[:max(limit, 1)]:
            rows.append({
                "contrato": record.contrato,
                "estado": record.estado,
                "cliente": record.cliente,
                "documento": format_document(record.documento),
                "fecha": record.fecha_operacion or record.fecha_inicio,
                "peso_total": format_decimal(record.peso_total_esperado),
                "peso_con_bolsa": format_decimal(record.peso_con_bolsa_esperado),
                "valor": format_money(record.valor),
            })
        return rows

    def finalize(self) -> dict[str, Any]:
        self._write_summary_sheets()
        self._format_audit_columns()
        self.save()

        cfg = load_config()
        cfg.update({
            "last_run": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "last_excel": str(self.excel_path) if self.excel_path else "",
            "last_mode": self.modo,
            "last_output": str(self.out_xlsx) if self.out_xlsx else "",
            "output_dir": str(self.output_base) if self.output_base else str(APP_DIR),
        })
        save_config(cfg)

        self._write_log_csv()
        progress = self.get_progress()
        progress["finalizado"] = True
        return progress

    def save(self) -> None:
        if self.wb and self.out_xlsx:
            self.wb.save(self.out_xlsx)

    # ---------------------------
    # Hojas de salida y logs
    # ---------------------------
    def _reset_sheet(self, name: str):
        if name in self.wb.sheetnames:
            sh = self.wb[name]
            sh.delete_rows(1, sh.max_row)
            return sh
        return self.wb.create_sheet(title=name)

    def _write_summary_sheets(self) -> None:
        progress = self.get_progress()
        resumen = self._reset_sheet("RESUMEN_CF")
        faltantes = self._reset_sheet("FALTANTES_CF")
        novedades = self._reset_sheet("NOVEDADES_CF")
        no_encontrados = self._reset_sheet("NO_ENCONTRADOS_CF")
        log_sheet = self._reset_sheet("LOG_CF")

        resumen_rows = [
            ["Campo", "Valor"],
            ["Fecha finalización", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
            ["Modo", self.modo],
            ["Fecha desde", progress["fecha_desde"]],
            ["Usuario/Auditor", self.usuario],
            ["Desarrollado por", DEVELOPER_SIGNATURE],
            ["Versión aplicación", APP_VERSION],
            ["Excel original", str(self.excel_path)],
            ["Excel salida", str(self.out_xlsx)],
            ["Peso bolsa seguridad usado", PESO_BOLSA_SEGURIDAD_GR],
            ["Tolerancia novedad peso", PESO_TOLERANCIA_GR],
            ["Contratos objetivo", progress["total_objetivo"]],
            ["Contratos verificados", progress["verificados"]],
            ["Contratos con novedad", progress["novedades"]],
            ["Contratos faltantes", progress["faltantes"]],
            ["Contratos no encontrados", progress["no_encontrados"]],
        ]
        for row in resumen_rows:
            resumen.append(row)

        faltantes.append(["Contrato", "Cliente", "Documento", "Valor", "Fecha base", "Cantidad Esperada", "Peso Libre", "Peso Total", "Peso con Bolsa", "Estado"])
        for record in sorted(self.target_contracts.values(), key=lambda x: min(x.rows) if x.rows else x.orden):
            if not self.is_verified(record):
                for row in record.rows:
                    for header in AUDIT_COLUMNS:
                        self.ws.cell(row=row, column=self.audit_cols[header]).fill = PatternFill("solid", fgColor="FFC7CE")
                faltantes.append([
                    record.contrato,
                    record.cliente,
                    format_document(record.documento),
                    format_money(record.valor),
                    record.fecha_operacion or record.fecha_inicio,
                    record.cantidad_esperada,
                    record.peso_libre_esperado,
                    record.peso_total_esperado,
                    record.peso_con_bolsa_esperado,
                    record.estado,
                ])

        novedades.append([
            "Contrato", "Cliente", "Documento", "Valor", "Novedad", "Observación", "Cantidad Esperada",
            "Peso Libre", "Peso Total", "Peso con Bolsa", "Fecha Auditoría"
        ])
        for record in sorted(self.target_contracts.values(), key=lambda x: min(x.rows) if x.rows else x.orden):
            if normalize_key(record.audit_estado) == "novedad":
                novedades.append([
                    record.contrato,
                    record.cliente,
                    format_document(record.documento),
                    format_money(record.valor),
                    record.audit_novedad_tipo,
                    record.audit_observacion,
                    record.cantidad_esperada,
                    record.peso_libre_esperado,
                    record.peso_total_esperado,
                    record.peso_con_bolsa_esperado,
                    record.audit_fecha,
                ])

        no_encontrados.append(["Contrato normalizado", "RAW escaneado", "Timestamp", "Detalle"])
        for contrato, info in sorted(self.no_encontrados.items()):
            no_encontrados.append([contrato, info.get("raw", ""), info.get("timestamp", ""), info.get("detalle", "")])

        log_sheet.append(["Timestamp", "RAW", "Contrato normalizado", "Resultado", "Detalle"])
        for row in self.audit_log:
            log_sheet.append(row)

        for sh in [resumen, faltantes, novedades, no_encontrados, log_sheet]:
            self._format_sheet(sh)

    def _format_sheet(self, sh) -> None:
        header_fill = PatternFill("solid", fgColor="1F4E78")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in sh[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for col in range(1, sh.max_column + 1):
            letter = get_column_letter(col)
            max_len = 10
            for row in range(1, min(sh.max_row, 200) + 1):
                value = sh.cell(row=row, column=col).value
                if value is not None:
                    max_len = max(max_len, min(len(str(value)), 60))
            sh.column_dimensions[letter].width = max_len + 2
        sh.freeze_panes = "A2"
        for row in sh.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _format_audit_columns(self) -> None:
        if not self.ws:
            return
        for header, col in self.audit_cols.items():
            letter = get_column_letter(col)
            self.ws.column_dimensions[letter].width = max(18, min(len(header) + 2, 34))
        self.ws.freeze_panes = self.ws.cell(row=self.header_row + 1, column=1).coordinate

    def _open_log_csv(self) -> None:
        if not self.log_csv:
            return
        self.log_csv.parent.mkdir(parents=True, exist_ok=True)
        if not self.log_csv.exists():
            with self.log_csv.open("w", newline="", encoding="utf-8-sig") as f:
                writer = csv.writer(f)
                writer.writerow(["timestamp", "raw", "normalizado", "resultado", "detalle"])

    def _append_log(self, row: list[Any]) -> None:
        self.audit_log.append(row)
        if self.log_csv:
            with self.log_csv.open("a", newline="", encoding="utf-8-sig") as f:
                csv.writer(f).writerow(row)

    def _write_log_csv(self) -> None:
        if not self.log_csv:
            return
        with self.log_csv.open("w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(["timestamp", "raw", "normalizado", "resultado", "detalle"])
            writer.writerows(self.audit_log)


__all__ = [
    "CajaFuerteBackend",
    "MODO_TOTAL",
    "MODO_PARCIAL",
    "get_last_run_text",
    "get_last_run_date",
    "get_default_output_dir",
    "parse_user_date",
    "normalize_contract",
    "PESO_BOLSA_SEGURIDAD_GR",
    "PESO_TOLERANCIA_GR",
]
