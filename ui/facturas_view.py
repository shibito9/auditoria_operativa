# -*- coding: utf-8 -*-
"""
Vista de Auditoría Facturas - Sistema de Auditoría Operativa
"""
import sys
from pathlib import Path

# Agregar directorio raíz al path para imports absolutos
ROOT_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(ROOT_DIR))

import customtkinter as ctk
from tkinter import filedialog, messagebox
from typing import Optional, Dict, Any
import pandas as pd
from datetime import date
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from services import CircularService, ValidacionComercialService


class FacturasView(ctk.CTkToplevel):
    def __init__(self, parent_window):
        super().__init__()
        
        self.parent_window = parent_window
        self.title("Auditoría Facturas - Sistema de Auditoría Operativa")
        self.geometry("1200x800")
        self.minsize(1000, 700)
        
        # Configuración de tema
        ctk.set_appearance_mode("Light")
        ctk.set_default_color_theme("blue")
        
        # Variables de estado
        self.excel_path: Optional[Path] = None
        self.df_facturas: Optional[pd.DataFrame] = None
        self.facturas_procesadas: Dict[str, Any] = {}
        
        # Servicios de validación comercial
        self.circular_service = CircularService()
        self.validacion_comercial = ValidacionComercialService(self.circular_service)
        
        # Configuración de la ventana
        self._setup_ui()
    
    def _setup_ui(self):
        """Configura la interfaz de usuario"""
        
        # Frame principal
        self.main_frame = ctk.CTkFrame(self)
        self.main_frame.pack(fill="both", expand=True, padx=20, pady=20)
        
        # Header
        self._create_header()
        
        # Panel de selección de archivo
        self._create_file_selection_panel()
        
        # Panel resumen
        self._create_summary_panel()
        
        # Tabla de resultados
        self._create_results_table()
        
        # Botones de acción
        self._create_action_buttons()
    
    def _create_header(self):
        """Crea el encabezado de la ventana"""
        header_frame = ctk.CTkFrame(self.main_frame)
        header_frame.pack(fill="x", pady=(0, 20))
        
        title_label = ctk.CTkLabel(
            header_frame,
            text="Auditoría de Facturas",
            font=ctk.CTkFont(size=24, weight="bold")
        )
        title_label.pack(side="left", padx=10, pady=10)
        
        subtitle_label = ctk.CTkLabel(
            header_frame,
            text="Validación de facturas con agrupación por Secuencia Factura",
            font=ctk.CTkFont(size=12)
        )
        subtitle_label.pack(side="left", padx=10, pady=10)
    
    def _create_file_selection_panel(self):
        """Crea el panel de selección de archivo Excel"""
        file_frame = ctk.CTkFrame(self.main_frame)
        file_frame.pack(fill="x", pady=(0, 20))
        
        # Label
        file_label = ctk.CTkLabel(
            file_frame,
            text="Archivo Excel de Facturas:",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        file_label.pack(side="left", padx=10, pady=10)
        
        # Entry para mostrar la ruta
        self.file_path_var = ctk.StringVar()
        file_entry = ctk.CTkEntry(
            file_frame,
            textvariable=self.file_path_var,
            width=400,
            font=ctk.CTkFont(size=12)
        )
        file_entry.pack(side="left", padx=10, pady=10)
        
        # Botón de búsqueda
        browse_button = ctk.CTkButton(
            file_frame,
            text="Buscar Excel",
            width=120,
            command=self._browse_excel
        )
        browse_button.pack(side="left", padx=10, pady=10)
        
        # Botón de carga
        load_button = ctk.CTkButton(
            file_frame,
            text="Cargar Excel",
            width=120,
            fg_color="#107C10",
            hover_color="#0B5A0A",
            command=self._load_and_validate_excel
        )
        load_button.pack(side="left", padx=10, pady=10)
        
        # Campo % máximo permitido
        self.porcentaje_max_var = ctk.StringVar(value="8")
        porcentaje_label = ctk.CTkLabel(
            file_frame,
            text="% máx permitido:",
            font=ctk.CTkFont(size=12)
        )
        porcentaje_label.pack(side="left", padx=(20, 5), pady=10)
        
        porcentaje_entry = ctk.CTkEntry(
            file_frame,
            textvariable=self.porcentaje_max_var,
            width=80,
            font=ctk.CTkFont(size=12)
        )
        porcentaje_entry.pack(side="left", padx=5, pady=10)
        
        # Campo tolerancia
        self.tolerancia_var = ctk.StringVar(value="0.10")
        tolerancia_label = ctk.CTkLabel(
            file_frame,
            text="Tolerancia:",
            font=ctk.CTkFont(size=12)
        )
        tolerancia_label.pack(side="left", padx=(20, 5), pady=10)
        
        tolerancia_entry = ctk.CTkEntry(
            file_frame,
            textvariable=self.tolerancia_var,
            width=80,
            font=ctk.CTkFont(size=12)
        )
        tolerancia_entry.pack(side="left", padx=5, pady=10)
        
        # Botón Validar Auditoría
        validate_button = ctk.CTkButton(
            file_frame,
            text="Validar Auditoría",
            width=140,
            fg_color="#1F4E78",
            hover_color="#163A5C",
            command=self._validate_auditoria
        )
        validate_button.pack(side="left", padx=10, pady=10)
    
    def _create_summary_panel(self):
        """Crea el panel resumen con estadísticas"""
        summary_frame = ctk.CTkFrame(self.main_frame)
        summary_frame.pack(fill="x", pady=(0, 20))
        
        # Título del panel
        summary_title = ctk.CTkLabel(
            summary_frame,
            text="Resumen de Auditoría",
            font=ctk.CTkFont(size=16, weight="bold")
        )
        summary_title.pack(pady=10)
        
        # Frame de métricas
        metrics_frame = ctk.CTkFrame(summary_frame)
        metrics_frame.pack(fill="x", padx=10, pady=10)
        
        # Métricas
        self._create_metric_card(
            metrics_frame, 
            "Total Facturas", 
            "0", 
            "total_facturas",
            "#1F4E78"
        )
        
        self._create_metric_card(
            metrics_frame, 
            "Correctas", 
            "0", 
            "correctas",
            "#107C10"
        )
        
        self._create_metric_card(
            metrics_frame, 
            "Errores", 
            "0", 
            "errores",
            "#C50F1F"
        )
        
        self._create_metric_card(
            metrics_frame, 
            "Revisión Manual", 
            "0", 
            "revision_manual",
            "#9A6700"
        )
        
        self._create_metric_card(
            metrics_frame, 
            "Excluidas", 
            "0", 
            "excluidas",
            "#555555"
        )
        
        self._create_metric_card(
            metrics_frame, 
            "Sin Regla", 
            "0", 
            "sin_regla",
            "#999999"
        )
        
        self._create_metric_card(
            metrics_frame, 
            "Fuera Vigencia", 
            "0", 
            "fuera_vigencia",
            "#FF9900"
        )
    
    def _create_metric_card(self, parent: ctk.CTkFrame, title: str, value: str, 
                           var_name: str, color: str):
        """Crea una tarjeta de métrica individual"""
        card_frame = ctk.CTkFrame(parent, fg_color=color)
        card_frame.pack(side="left", expand=True, fill="both", padx=5, pady=5)
        
        label_title = ctk.CTkLabel(
            card_frame,
            text=title,
            font=ctk.CTkFont(size=11, weight="bold"),
            text_color="white"
        )
        label_title.pack(pady=5)
        
        # Variable para el valor
        setattr(self, f"{var_name}_var", ctk.StringVar(value=value))
        
        label_value = ctk.CTkLabel(
            card_frame,
            textvariable=getattr(self, f"{var_name}_var"),
            font=ctk.CTkFont(size=20, weight="bold"),
            text_color="white"
        )
        label_value.pack(pady=5)
    
    def _create_results_table(self):
        """Crea la tabla de resultados"""
        table_frame = ctk.CTkFrame(self.main_frame)
        table_frame.pack(fill="both", expand=True, pady=(0, 20))
        
        # Título
        table_title = ctk.CTkLabel(
            table_frame,
            text="Resultados de Auditoría por Factura",
            font=ctk.CTkFont(size=14, weight="bold")
        )
        table_title.pack(pady=10)
        
        # Scrollable frame para la tabla
        self.table_scroll = ctk.CTkScrollableFrame(table_frame, height=400)
        self.table_scroll.pack(fill="both", expand=True, padx=10, pady=10)
        
        # Mensaje inicial
        self.empty_message = ctk.CTkLabel(
            self.table_scroll,
            text="Carga un archivo Excel para ver los resultados de auditoría",
            font=ctk.CTkFont(size=12),
            text_color="gray"
        )
        self.empty_message.pack(pady=50)
    
    def _create_action_buttons(self):
        """Crea los botones de acción"""
        buttons_frame = ctk.CTkFrame(self.main_frame)
        buttons_frame.pack(fill="x", pady=(0, 10))
        
        # Botón exportar
        export_button = ctk.CTkButton(
            buttons_frame,
            text="Exportar Resultados",
            width=150,
            command=self._export_results
        )
        export_button.pack(side="right", padx=10, pady=10)
        
        # Botón volver
        back_button = ctk.CTkButton(
            buttons_frame,
            text="Volver al Menú Principal",
            width=180,
            fg_color="#555555",
            hover_color="#444444",
            command=self._back_to_main
        )
        back_button.pack(side="right", padx=10, pady=10)
    
    def _browse_excel(self):
        """Abre el diálogo para seleccionar archivo Excel"""
        file_path = filedialog.askopenfilename(
            title="Selecciona el archivo Excel de Facturas",
            filetypes=[
                ("Excel files", "*.xlsx *.xls"),
                ("All files", "*.*")
            ]
        )
        
        if file_path:
            self.excel_path = Path(file_path)
            self.file_path_var.set(str(self.excel_path))
    
    def _load_and_validate_excel(self):
        """Carga y valida el archivo Excel"""
        if not self.excel_path or not self.excel_path.exists():
            messagebox.showerror(
                "Error",
                "Por favor selecciona un archivo Excel válido."
            )
            return
        
        try:
            # Cargar el Excel con pandas
            self.df_facturas = pd.read_excel(self.excel_path)
            
            # Validar columnas requeridas
            self._validate_required_columns()
            
            # Inicializar variables de resultados
            self.facturas_resultados = []
            
            messagebox.showinfo(
                "Éxito",
                f"Archivo cargado correctamente.\nTotal de filas: {len(self.df_facturas)}\n\n"
                "Ahora puedes configurar los parámetros y hacer clic en 'Validar Auditoría'."
            )
            
        except Exception as e:
            messagebox.showerror(
                "Error al cargar Excel",
                f"No se pudo cargar el archivo: {str(e)}"
            )
    
    def _validate_auditoria(self):
        """Ejecuta la validación de auditoría con los parámetros configurados"""
        if self.df_facturas is None:
            messagebox.showerror(
                "Error",
                "Primero debes cargar un archivo Excel."
            )
            return
        
        try:
            # Obtener parámetros
            porcentaje_permitido = float(self.porcentaje_max_var.get())
            tolerancia = float(self.tolerancia_var.get())
            
            # Procesar facturas con validación
            self.facturas_resultados = self._group_and_classify_facturas(
                self.df_facturas,
                porcentaje_permitido,
                tolerancia
            )
            
            # Actualizar interfaz
            self._update_summary()
            self._populate_results_table()
            
            messagebox.showinfo(
                "Validación Completada",
                f"Se validaron {len(self.facturas_resultados)} facturas.\n"
                f"Revisa el resumen y la tabla de resultados."
            )
            
        except ValueError as e:
            messagebox.showerror(
                "Error de parámetros",
                f"Por favor verifica los valores de % máximo y tolerancia: {e}"
            )
        except Exception as e:
            messagebox.showerror(
                "Error al validar",
                f"No se pudo completar la validación: {str(e)}"
            )
    
    def _validate_required_columns(self):
        """Valida que el Excel tenga las columnas requeridas con alias de encabezado"""
        # Definir columnas requeridas con sus alias permitidos
        required_columns_with_aliases = {
            "Secuencia Factura": ["Secuencia Factura", "Numero Factura", "Número Factura"],
            "Fecha Factura": ["Fecha Factura", "Fecha", "Fecha Factura"],
            "Factura Gramos": ["Factura Gramos", "Gramos Factura", "Peso Factura"],
            "Subtotal": ["Subtotal", "Sub Total"],
            "Descuento": ["Descuento", "Descuentos"],
            "Valor Total": ["Valor Total", "Total", "Valor Total"],
            "Producto Categoria": ["Producto Categoria", "Categoría Producto", "Categoria"],
            "Producto Sku": ["Producto Sku", "SKU", "Sku"],
            "Producto Gramos": ["Producto Gramos", "Gramos Producto", "Peso Producto"],
            "Producto Descripcion": ["Producto Descripcion", "Descripción Producto", "Descripcion"],
            "Linea Nombre": ["Linea Nombre", "Línea", "Linea"],
        }
        
        # Columna opcional para origen_comercial (se usará como fuente principal del cruce comercial)
        optional_columns_with_aliases = {
            "origen_comercial": [
                "Linea Nombre",  # Prioridad 1
                "Linea Origen",  # Prioridad 2
                "Procedencia",   # Prioridad 3
                "Procedencia Joya",  # Prioridad 4
                "Origen"         # Prioridad 5
            ]
        }
        
        # Mapeo de columnas encontradas en el Excel
        self.column_mapping = {}
        missing_columns = []
        
        for main_col, aliases in required_columns_with_aliases.items():
            found = False
            for df_col in self.df_facturas.columns:
                # Buscar coincidencia con cualquier alias (case-insensitive)
                for alias in aliases:
                    if alias.lower() in str(df_col).lower():
                        self.column_mapping[main_col] = df_col
                        found = True
                        break
                if found:
                    break
            
            if not found:
                missing_columns.append(f"{main_col} (alias: {', '.join(aliases)})")
        
        if missing_columns:
            raise ValueError(
                f"Faltan columnas requeridas en el Excel:\n{chr(10).join(missing_columns)}"
            )
        
        # Mapear origen_comercial (opcional, pero crítica para cruce comercial)
        self.origen_comercial_mapping = None
        for main_col, aliases in optional_columns_with_aliases.items():
            for df_col in self.df_facturas.columns:
                for alias in aliases:
                    if alias.lower() in str(df_col).lower():
                        self.origen_comercial_mapping = df_col
                        break
                if self.origen_comercial_mapping:
                    break
            if self.origen_comercial_mapping:
                break
        
        # Definir columna identificadora principal de factura
        self.factura_id_column = self.column_mapping["Secuencia Factura"]
    
    def _calculate_line_discount(self, row, porcentaje_permitido, tolerancia):
        """Calcula y clasifica una línea individual"""
        # Extraer valores del mapeo de columnas
        subtotal = row[self.column_mapping["Subtotal"]]
        descuento = row[self.column_mapping["Descuento"]]
        
        # Validar datos suficientes
        if pd.isna(subtotal) or pd.isna(descuento) or subtotal <= 0:
            return "DATOS_INSUFICIENTES", 0, "Subtotal vacío, Descuento vacío o Subtotal <= 0"
        
        # Calcular porcentaje real
        porcentaje_real = (descuento / subtotal) * 100
        
        # Clasificar según reglas V1
        limite_superior = porcentaje_permitido + tolerancia
        
        if descuento == 0:
            return "SIN_DESCUENTO", porcentaje_real, "Descuento es 0"
        elif porcentaje_real <= limite_superior:
            return "CORRECTA", porcentaje_real, f"{porcentaje_real:.2f}% <= {limite_superior:.2f}%"
        else:
            return "NOVEDAD", porcentaje_real, f"{porcentaje_real:.2f}% > {limite_superior:.2f}%"
    
    def _group_and_classify_facturas(self, df, porcentaje_permitido, tolerancia):
        """Agrupa por Secuencia Factura y clasifica cada factura con validación comercial"""
        # Calcular clasificación por línea (validación de descuento tradicional)
        lineas_clasificadas = []
        for idx, row in df.iterrows():
            estado, porcentaje, obs = self._calculate_line_discount(row, porcentaje_permitido, tolerancia)
            
            # Construir origen_comercial desde los aliases en orden de prioridad
            origen_comercial = ""
            if self.origen_comercial_mapping:
                valor_raw = row[self.origen_comercial_mapping]
                if pd.notna(valor_raw):
                    origen_comercial = str(valor_raw).strip().upper()
            
            lineas_clasificadas.append({
                "secuencia": row[self.factura_id_column],
                "estado": estado,
                "porcentaje": porcentaje,
                "observacion": obs,
                "subtotal": row[self.column_mapping["Subtotal"]],
                "descuento": row[self.column_mapping["Descuento"]],
                "fecha_factura": row[self.column_mapping["Fecha Factura"]],
                "categoria": row[self.column_mapping["Producto Categoria"]],
                "descripcion": row[self.column_mapping["Producto Descripcion"]],
                "linea": row[self.column_mapping["Linea Nombre"]],
                "gramos": row[self.column_mapping["Producto Gramos"]],
                "valor_total": row[self.column_mapping["Valor Total"]],
                "origen_comercial": origen_comercial
            })
        
        # Agrupar por Secuencia Factura
        facturas_agrupadas = {}
        for linea in lineas_clasificadas:
            secuencia = linea["secuencia"]
            if secuencia not in facturas_agrupadas:
                facturas_agrupadas[secuencia] = {
                    "secuencia": secuencia,
                    "cantidad_lineas": 0,
                    "subtotal_total": 0,
                    "descuento_total": 0,
                    "estados_lineas": [],
                    "observaciones": [],
                    "lineas_completas": []
                }
            
            facturas_agrupadas[secuencia]["cantidad_lineas"] += 1
            facturas_agrupadas[secuencia]["subtotal_total"] += linea["subtotal"]
            facturas_agrupadas[secuencia]["descuento_total"] += linea["descuento"]
            facturas_agrupadas[secuencia]["estados_lineas"].append(linea["estado"])
            facturas_agrupadas[secuencia]["observaciones"].append(linea["observacion"])
            facturas_agrupadas[secuencia]["lineas_completas"].append(linea)
        
        # Determinar estado final por factura (combinando validación tradicional + comercial)
        resultados = []
        for secuencia, datos in facturas_agrupadas.items():
            # Primero, validación tradicional de descuento
            estados_tradicionales = set(datos["estados_lineas"])
            
            if "DATOS_INSUFICIENTES" in estados_tradicionales:
                estado_tradicional = "REVISION_MANUAL"
            elif "NOVEDAD" in estados_tradicionales:
                estado_tradicional = "ERROR"
            else:
                estado_tradicional = "CORRECTA"
            
            # Segundo, validación comercial con circulares
            fecha_factura = datos["lineas_completas"][0]["fecha_factura"]
            
            # Convertir fecha si es necesario
            if isinstance(fecha_factura, str):
                try:
                    fecha_factura = pd.to_datetime(fecha_factura).date()
                except:
                    fecha_factura = date.today()
            
            # Preparar líneas para validación comercial
            lineas_para_validacion = []
            for linea in datos["lineas_completas"]:
                lineas_para_validacion.append({
                    "Producto Categoria": linea["categoria"],
                    "Producto Descripcion": linea["descripcion"],
                    "Linea Nombre": linea["linea"],
                    "Producto Gramos": linea["gramos"],
                    "Subtotal": linea["subtotal"],
                    "Descuento": linea["descuento"],
                    "Valor Total": linea["valor_total"],
                    "origen_comercial": linea.get("origen_comercial", "")
                })
            
            # Ejecutar validación comercial
            resultado_comercial = self.validacion_comercial.validar_factura(
                lineas_para_validacion,
                fecha_factura,
                tolerancia
            )
            
            # Separar estados explícitamente
            estado_descuento = estado_tradicional
            estado_comercial = resultado_comercial["estado"]
            
            # Matriz de combinación de estados
            if estado_descuento == "DATOS_INSUFICIENTES":
                estado_final = "REVISION_MANUAL"
            elif estado_descuento == "ERROR" and estado_comercial in ["CORRECTA", "SIN_REGLA", "FUERA_DE_VIGENCIA"]:
                estado_final = "ERROR_DESCUENTO"
            elif estado_descuento == "CORRECTA" and estado_comercial == "CORRECTA":
                estado_final = "CORRECTA"
            elif estado_descuento == "CORRECTA" and estado_comercial == "SIN_REGLA":
                estado_final = "SIN_REGLA"
            elif estado_descuento == "CORRECTA" and estado_comercial == "FUERA_DE_VIGENCIA":
                estado_final = "FUERA_DE_VIGENCIA"
            elif estado_descuento == "CORRECTA" and estado_comercial in ["ERROR_VALOR_VENTA", "ERROR_DESCUENTO"]:
                estado_final = estado_comercial
            elif estado_descuento == "ERROR" and estado_comercial in ["ERROR_VALOR_VENTA", "ERROR_DESCUENTO"]:
                estado_final = "ERROR_DESCUENTO_" + estado_comercial
            else:
                # Fallback: prioridad al estado comercial si es más severo
                estado_final = estado_comercial if estado_comercial != "SIN_REGLA" else estado_descuento
            
            # Construir observación combinada
            obs_descuento = ""
            if estado_descuento == "ERROR":
                obs_descuento = f"Error descuento: {datos['observaciones'][0] if datos['observaciones'] else ''}"
            elif estado_descuento == "DATOS_INSUFICIENTES":
                obs_descuento = "Datos insuficientes para validación"
            
            obs_comercial = resultado_comercial["observacion"]
            
            if obs_descuento and obs_comercial:
                observacion_final = f"{obs_descuento} | {obs_comercial}"
            elif obs_descuento:
                observacion_final = obs_descuento
            else:
                observacion_final = obs_comercial
            
            # Calcular % global
            porcentaje_global = (datos["descuento_total"] / datos["subtotal_total"]) * 100 if datos["subtotal_total"] > 0 else 0
            
            resultados.append({
                "secuencia": secuencia,
                "cantidad_lineas": datos["cantidad_lineas"],
                "subtotal_total": datos["subtotal_total"],
                "descuento_total": datos["descuento_total"],
                "porcentaje_global": porcentaje_global,
                "estado_descuento": estado_descuento,
                "estado_comercial": estado_comercial,
                "estado": estado_final,
                "circular_aplicada": resultado_comercial.get("circular_aplicada"),
                "observacion": observacion_final
            })
        
        return resultados
    
    def _update_summary(self):
        """Actualiza el panel resumen con las estadísticas"""
        if not hasattr(self, 'facturas_resultados') or not self.facturas_resultados:
            return
        
        # Calcular estadísticas
        total_facturas = len(self.facturas_resultados)
        correctas = sum(1 for f in self.facturas_resultados if f["estado"] == "CORRECTA")
        errores = sum(1 for f in self.facturas_resultados if f["estado"] in ["ERROR", "ERROR_DESCUENTO", "ERROR_VALOR_VENTA", "ERROR_DESCUENTO_ERROR_VALOR_VENTA", "ERROR_DESCUENTO_ERROR_DESCUENTO"])
        revision_manual = sum(1 for f in self.facturas_resultados if f["estado"] == "REVISION_MANUAL")
        excluidas = sum(1 for f in self.facturas_resultados if f["estado"] == "EXCLUIDA_POR_CIRCULAR")
        sin_regla = sum(1 for f in self.facturas_resultados if f["estado"] == "SIN_REGLA")
        fuera_vigencia = sum(1 for f in self.facturas_resultados if f["estado"] == "FUERA_DE_VIGENCIA")
        
        # Actualizar tarjetas
        self.total_facturas_var.set(str(total_facturas))
        self.correctas_var.set(str(correctas))
        self.errores_var.set(str(errores))
        self.revision_manual_var.set(str(revision_manual))
        self.excluidas_var.set(str(excluidas))
        self.sin_regla_var.set(str(sin_regla))
        self.fuera_vigencia_var.set(str(fuera_vigencia))
    
    def _populate_results_table(self):
        """Pobla la tabla de resultados con las facturas procesadas"""
        # Limpiar mensaje vacío
        if hasattr(self, 'empty_message'):
            self.empty_message.destroy()
        
        # Limpiar contenido anterior
        for widget in self.table_scroll.winfo_children():
            widget.destroy()
        
        if not hasattr(self, 'facturas_resultados') or not self.facturas_resultados:
            placeholder_label = ctk.CTkLabel(
                self.table_scroll,
                text="Valida la auditoría para ver los resultados",
                font=ctk.CTkFont(size=12),
                text_color="gray"
            )
            placeholder_label.pack(pady=50)
            return
        
        # Crear encabezado de tabla
        header_frame = ctk.CTkFrame(self.table_scroll)
        header_frame.pack(fill="x", padx=5, pady=5)
        
        headers = ["Secuencia", "Líneas", "Subtotal", "Descuento", "% Global", "Estado Descuento", "Estado Comercial", "Estado Final", "Circular", "Observación"]
        for i, header in enumerate(headers):
            label = ctk.CTkLabel(
                header_frame,
                text=header,
                font=ctk.CTkFont(size=11, weight="bold"),
                width=100 if i < 8 else 120,
                anchor="w"
            )
            label.grid(row=0, column=i, padx=5, pady=5)
        
        # Poblar filas con resultados
        for i, factura in enumerate(self.facturas_resultados, start=1):
            row_frame = ctk.CTkFrame(self.table_scroll)
            row_frame.pack(fill="x", padx=5, pady=2)
            
            # Color según estado final
            bg_color = {
                "CORRECTA": "#C6EFCE",
                "ERROR": "#FFC7CE",
                "ERROR_DESCUENTO": "#FFC7CE",
                "ERROR_VALOR_VENTA": "#FFC7CE",
                "ERROR_DESCUENTO_ERROR_VALOR_VENTA": "#FFC7CE",
                "ERROR_DESCUENTO_ERROR_DESCUENTO": "#FFC7CE",
                "REVISION_MANUAL": "#FFF2CC",
                "FUERA_DE_VIGENCIA": "#FFE6CC",
                "EXCLUIDA_POR_CIRCULAR": "#E2E2E2",
                "REQUIERE_OTRA_CIRCULAR": "#D9E1F2",
                "SIN_REGLA": "#F2F2F2"
            }.get(factura["estado"], "#FFFFFF")
            
            # Colores para estados individuales
            bg_descuento = {
                "CORRECTA": "#C6EFCE",
                "ERROR": "#FFC7CE",
                "DATOS_INSUFICIENTES": "#FFF2CC"
            }.get(factura.get("estado_descuento", ""), "#FFFFFF")
            
            bg_comercial = {
                "CORRECTA": "#C6EFCE",
                "SIN_REGLA": "#F2F2F2",
                "FUERA_DE_VIGENCIA": "#FFE6CC",
                "ERROR_VALOR_VENTA": "#FFC7CE",
                "ERROR_DESCUENTO": "#FFC7CE",
                "EXCLUIDA_POR_CIRCULAR": "#E2E2E2",
                "REQUIERE_OTRA_CIRCULAR": "#D9E1F2",
                "REVISION_MANUAL": "#FFF2CC"
            }.get(factura.get("estado_comercial", ""), "#FFFFFF")
            
            # Secuencia
            ctk.CTkLabel(
                row_frame,
                text=str(factura["secuencia"]),
                width=100,
                anchor="w"
            ).grid(row=0, column=0, padx=5, pady=5)
            
            # Cantidad líneas
            ctk.CTkLabel(
                row_frame,
                text=str(factura["cantidad_lineas"]),
                width=100,
                anchor="w"
            ).grid(row=0, column=1, padx=5, pady=5)
            
            # Subtotal total
            ctk.CTkLabel(
                row_frame,
                text=f"${factura['subtotal_total']:,.2f}",
                width=100,
                anchor="w"
            ).grid(row=0, column=2, padx=5, pady=5)
            
            # Descuento total
            ctk.CTkLabel(
                row_frame,
                text=f"${factura['descuento_total']:,.2f}",
                width=100,
                anchor="w"
            ).grid(row=0, column=3, padx=5, pady=5)
            
            # % global
            ctk.CTkLabel(
                row_frame,
                text=f"{factura['porcentaje_global']:.2f}%",
                width=100,
                anchor="w"
            ).grid(row=0, column=4, padx=5, pady=5)
            
            # Estado Descuento
            estado_descuento_label = ctk.CTkLabel(
                row_frame,
                text=factura.get("estado_descuento", ""),
                width=100,
                anchor="w",
                fg_color=bg_descuento,
                corner_radius=4
            )
            estado_descuento_label.grid(row=0, column=5, padx=5, pady=5)
            
            # Estado Comercial
            estado_comercial_label = ctk.CTkLabel(
                row_frame,
                text=factura.get("estado_comercial", ""),
                width=100,
                anchor="w",
                fg_color=bg_comercial,
                corner_radius=4
            )
            estado_comercial_label.grid(row=0, column=6, padx=5, pady=5)
            
            # Estado Final
            estado_final_label = ctk.CTkLabel(
                row_frame,
                text=factura["estado"],
                width=100,
                anchor="w",
                fg_color=bg_color,
                corner_radius=4
            )
            estado_final_label.grid(row=0, column=7, padx=5, pady=5)
            
            # Circular aplicada
            ctk.CTkLabel(
                row_frame,
                text=factura.get("circular_aplicada", "-"),
                width=120,
                anchor="w"
            ).grid(row=0, column=8, padx=5, pady=5)
            
            # Observación
            obs_text = factura["observacion"]
            obs_display = obs_text[:40] + "..." if len(obs_text) > 40 else obs_text
            ctk.CTkLabel(
                row_frame,
                text=obs_display,
                width=200,
                anchor="w"
            ).grid(row=0, column=9, padx=5, pady=5)
    
    def _export_results(self):
        """Exporta los resultados a un archivo Excel"""
        if not hasattr(self, 'facturas_resultados') or not self.facturas_resultados:
            messagebox.showwarning("Advertencia", "No hay resultados para exportar")
            return
        
        try:
            # Solicitar ruta de guardado
            output_dir = ROOT_DIR / "Resultados"
            output_dir.mkdir(exist_ok=True)
            
            default_filename = f"auditoria_facturas_{date.today().strftime('%Y%m%d_%H%M')}.xlsx"
            
            file_path = filedialog.asksaveasfilename(
                title="Guardar resultados de auditoría",
                defaultextension=".xlsx",
                initialdir=str(output_dir),
                initialfile=default_filename,
                filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
            )
            
            if not file_path:
                return  # Usuario canceló
            
            # Preparar datos para exportación usando self.facturas_resultados
            export_data = []
            for factura in self.facturas_resultados:
                export_data.append({
                    "Secuencia": factura["secuencia"],
                    "Lineas": factura["cantidad_lineas"],
                    "Subtotal": factura["subtotal_total"],
                    "Descuento": factura["descuento_total"],
                    "% Global": factura["porcentaje_global"],
                    "Estado Descuento": factura.get("estado_descuento", ""),
                    "Estado Comercial": factura.get("estado_comercial", ""),
                    "Estado Final": factura["estado"],
                    "Circular": factura.get("circular_aplicada", ""),
                    "Observacion": factura["observacion"]
                })
            
            # Crear DataFrame y exportar
            df_export = pd.DataFrame(export_data)
            df_export.to_excel(file_path, index=False, engine='openpyxl', sheet_name='Resultados')
            
            # Aplicar formato visual con openpyxl
            wb = load_workbook(file_path)
            ws = wb['Resultados']
            
            # Estilo para encabezados
            header_font = Font(bold=True, size=11)
            header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar estilo a encabezados
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
            
            # Activar autofiltro
            ws.auto_filter.ref = ws.dimensions
            
            # Congelar fila superior
            ws.freeze_panes = "A2"
            
            # Formatear columnas numéricas
            from openpyxl.styles import numbers
            
            # Encontrar columnas por nombre
            header_row = 1
            col_indices = {}
            for col in range(1, ws.max_column + 1):
                cell_value = ws.cell(row=header_row, column=col).value
                if cell_value:
                    col_indices[str(cell_value)] = col
            
            # Formatear Subtotal y Descuento (separador de miles, 2 decimales)
            if "Subtotal" in col_indices:
                col = col_indices["Subtotal"]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
            
            if "Descuento" in col_indices:
                col = col_indices["Descuento"]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).number_format = '#,##0.00'
            
            # Formatear % Global (2 decimales con símbolo %)
            if "% Global" in col_indices:
                col = col_indices["% Global"]
                for row in range(2, ws.max_row + 1):
                    ws.cell(row=row, column=col).number_format = '0.00%'
            
            # Ajustar ancho automático de columnas
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            # Crear hoja de Resumen
            if self.facturas_resultados:
                ws_resumen = wb.create_sheet("Resumen")
                
                # Calcular estadísticas
                total_facturas = len(self.facturas_resultados)
                correctas = sum(1 for f in self.facturas_resultados if f["estado"] == "CORRECTA")
                errores = sum(1 for f in self.facturas_resultados if f["estado"] in ["ERROR", "ERROR_DESCUENTO", "ERROR_VALOR_VENTA", "ERROR_DESCUENTO_ERROR_VALOR_VENTA", "ERROR_DESCUENTO_ERROR_DESCUENTO"])
                revision_manual = sum(1 for f in self.facturas_resultados if f["estado"] == "REVISION_MANUAL")
                sin_regla = sum(1 for f in self.facturas_resultados if f["estado"] == "SIN_REGLA")
                fuera_vigencia = sum(1 for f in self.facturas_resultados if f["estado"] == "FUERA_DE_VIGENCIA")
                
                subtotal_total = sum(f["subtotal_total"] for f in self.facturas_resultados)
                descuento_total = sum(f["descuento_total"] for f in self.facturas_resultados)
                
                # Escribir resumen
                ws_resumen['A1'] = "RESUMEN DE AUDITORÍA"
                ws_resumen['A1'].font = Font(bold=True, size=14)
                ws_resumen.merge_cells('A1:B1')
                
                ws_resumen['A3'] = "Métrica"
                ws_resumen['B3'] = "Valor"
                ws_resumen['A3'].font = Font(bold=True)
                ws_resumen['B3'].font = Font(bold=True)
                
                resumen_data = [
                    ("Total Facturas", total_facturas),
                    ("Correctas", correctas),
                    ("Errores", errores),
                    ("Revisión Manual", revision_manual),
                    ("Sin Regla", sin_regla),
                    ("Fuera de Vigencia", fuera_vigencia),
                    ("", ""),
                    ("Subtotal Total", subtotal_total),
                    ("Descuento Total", descuento_total),
                    ("% Global Promedio", (descuento_total / subtotal_total * 100) if subtotal_total > 0 else 0)
                ]
                
                for i, (metrica, valor) in enumerate(resumen_data, start=4):
                    ws_resumen[f'A{i}'] = metrica
                    ws_resumen[f'B{i}'] = valor
                    
                    # Formatear valores numéricos
                    if isinstance(valor, (int, float)) and metrica not in ["Total Facturas", "Correctas", "Errores", "Revisión Manual", "Sin Regla", "Fuera de Vigencia"]:
                        if "%" in metrica:
                            ws_resumen[f'B{i}'].number_format = '0.00%'
                        else:
                            ws_resumen[f'B{i}'].number_format = '#,##0.00'
                
                # Ajustar ancho de columnas en resumen
                ws_resumen.column_dimensions['A'].width = 25
                ws_resumen.column_dimensions['B'].width = 20
            
            # Guardar cambios
            wb.save(file_path)
            
            messagebox.showinfo(
                "Éxito",
                f"Resultados exportados correctamente a:\n{file_path}\n\nHoja 'Resultados': detalle de facturas\nHoja 'Resumen': estadísticas generales"
            )
            
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudo exportar los resultados:\n{str(e)}"
            )
    
    def _back_to_main(self):
        """Vuelve a la ventana principal"""
        self.destroy()
        if self.parent_window:
            self.parent_window.deiconify()
